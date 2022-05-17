__version__ = '1.0'
__date__ = '11/05/2022'
__author__ = 'Juan Ledesma'

import os
import shutil
import pandas as pd
import numpy as np
import argparse
from Bio import SeqIO, AlignIO
from Bio.Align.Applications import ClustalwCommandline
from Bio.Align import MultipleSeqAlignment
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import mutation_utilities as mu

program_specs = argparse.ArgumentParser(
    prog='hbv_mutation_genotyping.py',
    usage='%(prog)s -dir <RUN location>',
    description='The script %(prog)s takes as an argument the RUN location\
                where a folder "Consensus" is, which contains the FASTA\
                sequences to be analysed and a csv file with their genotyping\
                results.\
                It returns an excel report with the mutations found for each\
                sequence, a log with details about the sequences analysed and\
                alignmnets for the user to inspect if needed.')
program_specs.add_argument(
    '-dir', 
    help='path to the directory where the fasta files are\
          This directory MUST contain a csv file with the genotyping results',
    required=True)

args = program_specs.parse_args()

try:
    dir_location = Path.cwd().joinpath(args.dir) 
    # direct access to refseq sequences. 
    # hbv_mutational_analysis OR HBV Routine Sequencing when deployed
    os.chdir(dir_location.parents[2])
    path_to_run = Path.cwd().joinpath(*dir_location.parts[-3:]) 
    path_to_input = path_to_run.joinpath('Consensus')
    run_name = path_to_run.name 
    if path_to_run.exists():
        if path_to_input.is_dir: 
            path_to_alns = path_to_input.joinpath('alignments') 
            path_to_trimmed = path_to_alns.joinpath('raw_trimmed') 
            path_to_tmp =  path_to_input.joinpath('tmp') 
            if path_to_tmp.exists():
                shutil.rmtree(path_to_tmp) 
            if path_to_alns.exists():
                shutil.rmtree(path_to_alns)
                os.mkdir(path_to_alns)
                os.mkdir(path_to_trimmed) # for checking not inframe sequences
                os.mkdir(path_to_tmp) #used for temp files
            else:
                os.mkdir(path_to_alns)
                os.mkdir(path_to_trimmed) 
                os.mkdir(path_to_tmp) 

        xprecore_refseq = SeqIO.read(
                        'refseqs/LC360507.1_WG_xprecore.fas','fasta')
        spol_refseqs = {'A':'AY128092.1_Genotype_A_polymerase_surface', 
                        'B':'AY167089.1_Genotype_B_polymerase_surface',
                        'C':'KM999990.1_Genotype_C_polymerase_surface',
                        'D':'X65257.1_Genotype_D_polymerase_surface',
                        'E':'X75657.1_Genotype_E_polymerase_surface',
                        'F':'KX264497.1_Genotype_F_polymerase_surface',
                        'G':'AF160501.1_Genotype_G_polymerase_surface',
                        'H':'AY090457.1_Genotype_H_polymerase_surface',
                        'I':'AF241411.1_Genotype_I_polymerase_surface',
                        'J':'AB486012.1_Genotype_J_polymerase_surface'
                        }
        try:
            csv_file = path_to_input.joinpath('Sample_list.csv') 
            genotype_results = pd.read_csv(csv_file)
            genotype_results = genotype_results.fillna("NA")

            if genotype_results.iloc[0]['Seq ID'] and\
                genotype_results.iloc[0]['Genotype']:    
                key = 0
                id_map = {} # to prevent issues with long IDs for the sequences
                report = {}
                xc_fasta_input = [] # used for summary log
                sp_fasta_input = [] # used for summary log
                all_fasta = []
                csv_seq_list = []
                not_labelled_seqs = []

                ''' (A) PAIR XC/SP CONSENSUS SEQUENCE TO SPECIFIC REFSEQ'''

                for fasta_file in path_to_input.glob('*.fas'):            
                    sequence= SeqIO.read(fasta_file, "fasta")
                    all_fasta.append(sequence.id)
                    name = (sequence.id)[:-2]
                    sequence.seq = sequence.seq.ungap()
                    key +=  1
                    id_map[str(key)] = name
                    report[name] = {'surface':{}, 
                                    'polymeraseRT':{}, 
                                    'xregion':{}, 
                                    'precore':{}
                                    }    

                    if 'xc' in sequence.id[-2:]: #xc/sp must be at the end of the id
                        xc_fasta_input.append(sequence.id)
                        xc_SeqRecord = []
                        xc_SeqRecord.append(xprecore_refseq)
                        sequence.id = str(key)
                        xc_SeqRecord.append(sequence)
                        output_name = path_to_tmp.joinpath(
                                            f'{name}_xc_refseq.fas')
                        SeqIO.write(xc_SeqRecord, output_name,'fasta')

                    elif 'sp' in sequence.id[-2:]: 
                        sp_fasta_input.append(sequence.id)
                        for row in range(len(genotype_results)):
                            csv_seq_list.append(
                                        genotype_results.iloc[row]['Seq ID']) 
                                                # list of IDs in the csv

                            ''' Selection of sur/pol refseqs based on genotype'''
                            if name in genotype_results.iloc[row]['Seq ID']:
                                genotype = genotype_results.iloc[row]['Genotype']                  
                                if genotype[0] in spol_refseqs:
                                    RefSeq = Path('refseqs').joinpath(
                                                f'{spol_refseqs[genotype[0]]}.fas')
                                    spol_refseq = SeqIO.read(RefSeq,'fasta')
                                    spol_SeqRecord = []
                                    spol_SeqRecord.append(spol_refseq)
                                    sequence.id = str(key)
                                    spol_SeqRecord.append(sequence)
                                    output_name = path_to_tmp.joinpath(
                                                    f'{name}_sp_refseq.fas')
                                    SeqIO.write(spol_SeqRecord, output_name,
                                                'fasta') 

                    else: # spol, xcore, any other thing, missing label
                        not_labelled_seqs.append(sequence.id)

            ''' (B) ALL THE SEQUENCES MUST HAVE THE LABEL xc/sp IN THE FASTA 
            HEADER TO INDENTIFY THE AMPLICONS TO ANALYSE. OTHERWISE AND LOG 
            ERROR IS GENERATED.'''

            if len(not_labelled_seqs) != 0:
                log_name = mu.error_log(path_to_run, run_name, not_labelled_seqs)
                print(f'''
    The analysis HAS BEEN ABORTED.

    FASTA sequences may not include the labels xc or sp to identify the PCR
    amplicon they were generated from (x region/core or surface/polymerase,
    respectively). 

    Please see the log 
    {log_name} 
    for additional details.
    ''') 


            ''' (C) INITIAL PAIR ALIGNMENT BETWEEN CONSENSUS SEQUENCES AND 
            REFSEQ, JUST TO TRIM THE AMPLICONS AT THE PROPER LENGTH'''

            if len(not_labelled_seqs) == 0:
                for fasta in path_to_tmp.glob('*refseq.fas'): 
                    #clustalw_exe = "/usr/bin/clustalw" # Windows users
                    #clustalw_cline = ClustalwCommandline(clustalw_exe, infile=str(fasta))
                    #assert os.path.isfile(clustalw_exe), "Clustal W executable missing"
                    clustalw_cline = ClustalwCommandline('clustalw2', 
                                                          infile=str(fasta))
                    stdout, stderr = clustalw_cline()

                # object  <class 'pathlib.PosixPath'> to string prevents from 
                # having this error. Spaces in the path? 
                # Bio.Application.ApplicationError: Non-zero return code 1 from 
                # 'clustalw2 -infile=/...hbv_mutational_analysis/2022 Sequencing Data
                # /Jan 2022/E1673/Consensus/tmp/09_H215100086_xc_refseq.fas', message 
                # 'Error: unknown option -equencing' 

                ''' (D) TRIMMING ACCORDING TO GENES, AND MUTATION ANALYSIS'''

                for aln in path_to_tmp.glob('*refseq.aln'): 
                
                    ''' Analysis of xc amplicons '''
                    # <class 'pathlib.PosixPath'> to string to iteration
                    if 'xc' in str(aln): 
                        xcore_aln_SeqRecords = AlignIO.read(aln, 'clustal')
                        xcore_aln_SeqRecords[1].id = id_map[str(xcore_aln_SeqRecords[1].id)]  
                        x_aln = mu.trimming_alignment(xcore_aln_SeqRecords, 
                                                      1374-1, 1838)
                        x_aln[0].id = str(x_aln[0].id).replace('WG','xregion') 
                        x_aln[0].description = ''
                        x_aln[1].description = 'x_region'
                        x_alignment = MultipleSeqAlignment(x_aln) 
                        output_name = path_to_trimmed.joinpath(
                                        f'{x_aln[1].id}_x_region_raw.fas')
                        AlignIO.write(x_alignment, output_name,'fasta')

                        precore_aln = mu.trimming_alignment(
                                        xcore_aln_SeqRecords, 1814-1, 1900)
                        precore_aln[0].id = str(precore_aln[0].id).replace('WG','precore') 
                        precore_aln[0].description = ''
                        precore_aln[1].description = 'precore'
                        precore_alignment = MultipleSeqAlignment(precore_aln)
                        output_name = path_to_trimmed.joinpath( 
                                        f'{precore_aln[1].id}_precore_raw.fas')
                        AlignIO.write(precore_alignment, output_name,'fasta')

                        if precore_aln:
                            sampleid = precore_aln[1].id 
                            region = (precore_aln[0].id).split('_')[-1]
                            refseqid = precore_aln[0].id

                            starting_aa,\
                            starting_nt = mu.first_aa_nt_in_partial_fragment(precore_aln)
                            frame = mu.frameshift_check(precore_aln, starting_nt)
                            report[sampleid][region]['frameshift'] = frame

                            if len(frame) == 0:
                                codonAlnRecord = mu.codon_alignment(precore_aln,
                                                                    region, 
                                                                    path_to_alns, 
                                                                    path_to_tmp, 
                                                                    starting_nt)

                                codons, aas = mu.translation_of_codons(codonAlnRecord)

                                insertions,\
                                codonsNumb, aasNumb = mu.refseq_numbering(codons,
                                                                          aas, 
                                                                          starting_aa)
                                report[sampleid][region]['insertions'] = insertions

                                mutations = mu.scan_mutations(codonsNumb, 
                                                              aasNumb, 
                                                              refseqid)
                                report[sampleid][region]['mutations'] = mutations  

                        if x_aln:
                            sampleid = (x_aln[1].id)
                            region = (x_aln[0].id).split('_')[-1]
                            refseqid = x_aln[0].id

                            ''' only final region for this gene, partial amplicon'''
                            starting_aa,\
                            starting_nt = mu.first_aa_nt_in_partial_fragment(x_aln)
                            frame = mu.frameshift_check(x_aln, starting_nt)
                            report[sampleid][region]['frameshift'] = frame

                            if len(frame) == 0:
                                codonAlnRecord = mu.codon_alignment(x_aln,
                                                                    region, 
                                                                    path_to_alns, 
                                                                    path_to_tmp, 
                                                                    starting_nt)

                                codons, aas = mu.translation_of_codons(codonAlnRecord)

                                insertions,\
                                codonsNumb, aasNumb = mu.refseq_numbering(codons, 
                                                                          aas, 
                                                                          starting_aa)                                            
                                report[sampleid][region]['insertions'] = insertions

                                mutations = mu.scan_mutations(codonsNumb, 
                                                              aasNumb, 
                                                              refseqid)

                                report[sampleid][region]['mutations'] = mutations         

                    ''' Analysis of sp amplicons '''
                    if 'sp' in str(aln): 
                        genotype_refseqs = { # 2076 (Whole RT) 1851 (to cover position 269)
                                'AY128092.1_Geno_A_polymerase':{
                                    'surface': [1070-1,1750],'rt':[1045-1,1851]
                                    },                                  #2076]}
                                'AY167089.1_Geno_B_polymerase':{
                                    'surface': [1064-1,1744], 'rt':[1039-1,1845]
                                    },                                  #2070]},
                                'KM999990.1_Geno_C_polymerase':{
                                    'surface': [1064-1,1744], 'rt':[1039-1,1845]
                                    },                                  #2070]},
                                'X65257.1_Geno_D_polymerase':{
                                    'surface': [1031-1,1711], 'rt':[1006-1,1812]
                                    },                                  #2037]}
                                'X75657.1_Geno_E_polymerase':{
                                    'surface': [1061-1,1741], 'rt':[1036-1,1842]
                                    },                                  #2067]}, 
                                'KX264497.1_Geno_F_polymerase':{
                                    'surface': [1064-1,1744], 'rt':[1039-1,1845]
                                    },                                  #2070]},
                                'AF160501.1_Geno_G_polymerase':{
                                    'surface': [1061-1,1741], 'rt':[1036-1,1842]
                                    },                                  #2067]} 
                                'AY090457.1_Geno_H_polymerase':{
                                    'surface': [1064-1,1744], 'rt':[1039-1,1845]
                                    },                                  #2070]},
                                'AF241411.1_Geno_I_polymerase':{
                                    'surface': [1064-1,1744], 'rt':[1039-1,1845]
                                    },                                  #2070]} 
                                'AB486012.1_Geno_J_polymerase':{
                                    'surface': [1031-1,1711], 'rt':[1006-1,1812]
                                    }                                   #2037]}  
                                            }

                        spol_aln_SeqRecords = AlignIO.read(aln, 'clustal')
                        spol_aln_SeqRecords[1].id = id_map[str(spol_aln_SeqRecords[1].id)]

                        if spol_aln_SeqRecords[0].id in genotype_refseqs:
                        
                            start_surface = genotype_refseqs[spol_aln_SeqRecords[0].id]['surface'][0]
                            end_surface = genotype_refseqs[spol_aln_SeqRecords[0].id]['surface'][1]
                            surface_aln = mu.trimming_alignment(spol_aln_SeqRecords,
                                                                start_surface, 
                                                                end_surface)
                            surface_aln[0].id = str(surface_aln[0].id).replace('polymerase', 
                                                                                'surface') 
                            surface_aln[0].description = ''
                            surface_aln[1].description = 'surface'
                            surface_alignment = MultipleSeqAlignment(surface_aln)
                            output_name = path_to_trimmed.joinpath(
                                            f'{surface_aln[1].id}_surface_raw.fas')
                            AlignIO.write(surface_alignment, output_name,'fasta')

                            start_rt = genotype_refseqs[spol_aln_SeqRecords[0].id]['rt'][0]
                            end_rt = genotype_refseqs[spol_aln_SeqRecords[0].id]['rt'][1]
                            pol_RT_aln = mu.trimming_alignment(spol_aln_SeqRecords, 
                                                               start_rt, 
                                                               end_rt)
                            pol_RT_aln[0].id = str(pol_RT_aln[0].id).replace('polymerase', 
                                                                            'polymeraseRT')
                            pol_RT_aln[0].description = ''
                            pol_RT_aln[1].description = 'pol RT'
                            polrt_alignment = MultipleSeqAlignment(pol_RT_aln)
                            output_name = path_to_trimmed.joinpath(
                                            f'{pol_RT_aln[1].id}_polymeraseRT_raw.fas')
                            AlignIO.write(polrt_alignment, output_name,'fasta')

                            if surface_aln:
                                sampleid = (surface_aln[1].id)
                                region = (surface_aln[0].id).split('_')[-1]
                                refseqid = surface_aln[0].id
                            
                                starting_aa,\
                                starting_nt = mu.first_aa_nt_in_partial_fragment(surface_aln)
                                frame = mu.frameshift_check(surface_aln, starting_nt)
                                report[sampleid][region]['frameshift'] = frame

                                if len(frame) == 0:
                                    codonAlnRecord = mu.codon_alignment(surface_aln,
                                                                        region, 
                                                                        path_to_alns, 
                                                                        path_to_tmp, 
                                                                        starting_nt)

                                    codons, aas = mu.translation_of_codons(codonAlnRecord)

                                    insertions,\
                                    codonsNumb, aasNumb = mu.refseq_numbering(codons, 
                                                                              aas, 
                                                                              starting_aa)
                                    report[sampleid][region]['insertions'] = insertions

                                    mutations = mu.scan_mutations(codonsNumb, 
                                                                  aasNumb, 
                                                                  refseqid)
                                    report[sampleid][region]['mutations'] = mutations  
                                
                            if pol_RT_aln:
                                sampleid = (pol_RT_aln[1].id)
                                region = (pol_RT_aln[0].id).split('_')[-1] # polymeraseRT
                                refseqid = pol_RT_aln[0].id
                                
                                starting_aa,\
                                starting_nt = mu.first_aa_nt_in_partial_fragment(pol_RT_aln)
                                frame = mu.frameshift_check(pol_RT_aln, starting_nt)
                                report[sampleid][region]['frameshift'] = frame

                                if len(frame) == 0:
                                    codonAlnRecord = mu.codon_alignment(pol_RT_aln,
                                                                        region, 
                                                                        path_to_alns, 
                                                                        path_to_tmp, 
                                                                        starting_nt)

                                    codons, aas = mu.translation_of_codons(codonAlnRecord)

                                    insertions,\
                                    codonsNumb, aasNumb = mu.refseq_numbering(codons, 
                                                                              aas, 
                                                                              starting_aa)
                                    report[sampleid][region]['insertions'] = insertions

                                    mutations = mu.scan_mutations(codonsNumb, 
                                                                  aasNumb, 
                                                                  refseqid)
                                    report[sampleid][region]['mutations'] = mutations  
 
                '''(E) CREATION OF AN EXCEL REPORT WITH RESULTS'''

                df_results = pd.DataFrame()
                sortedreport = dict(sorted(report.items()))
                row = {'Notes':''}

                csv_molis_ids = [gr.split('_')[1] for gr in genotype_results['Seq ID']]
                csv_tube_numbers = [gr.split('_')[0] for gr in genotype_results['Seq ID']]
                genotype_results['Molis No.'] = csv_molis_ids 
                genotype_results['Tube No.'] = csv_tube_numbers

                for seqid, geneScan in sortedreport.items():
                    row['Tube No.'] = seqid.split('_')[0]
                    row['Molis No.'] = seqid.split('_')[1]  
                    for gene in geneScan:
                        if geneScan[gene] != {}: 
                            if len(geneScan[gene]['frameshift']) != 0:
                                info =', '.join(geneScan[gene]['frameshift'])
                                row[gene] = info
                            else:
                                if len(geneScan[gene]['insertions']) == 0 and \
                                    len(geneScan[gene]['mutations']) == 0:
                                    row[gene] = "No mutations"
                                if len(geneScan[gene]['insertions']) == 0 and \
                                    len(geneScan[gene]['mutations']) != 0:
                                    info =', '.join(geneScan[gene]['mutations'])
                                    row[gene] = info
                                if len(geneScan[gene]['insertions']) != 0 and \
                                    len(geneScan[gene]['mutations']) != 0:
                                    ins_mut = geneScan[gene]['insertions'] +\
                                              geneScan[gene]['mutations']
                                    info =', '.join(ins_mut)
                                    row[gene] = info
                        else:
                            row[gene] = 'NA'
                    df_results = df_results.append(row, ignore_index = True)

                combined_df_results = pd.merge(genotype_results, df_results, 
                                               how='left',
                                               on=['Molis No.','Tube No.']) 
                combined_df_results = combined_df_results.replace({np.nan: 'NA' })
                #combined_df_results = combined_df_results.fillna("NA")
                combined_df_results['Notes']= ''
                combined_df_results.rename(columns={'surface':'Surface',
                                                    'polymeraseRT': 'Polymerase (RT)', 
                                                    'xregion': 'X region', 
                                                    'precore': 'Precore' }, 
                                           inplace= True)
                combined_df_results = combined_df_results[['Tube No.', 
                                                           'Molis No.', 
                                                           'Genotype', 
                                                           'Surface',
                                                           'Polymerase (RT)', 
                                                           'X region', 'Precore', 
                                                           'Notes']]

                workbook = Workbook()
                ws = workbook.active
                ws1 = workbook.create_sheet("All", 0)
                ws2 = workbook.create_sheet("A", 1)
                ws3 = workbook.create_sheet("B", 2)
                ws4 = workbook.create_sheet("C", 3)
                ws5 = workbook.create_sheet("D", 4)
                ws6 = workbook.create_sheet("E", 5)
                ws7 = workbook.create_sheet("F", 6)
                ws8 = workbook.create_sheet("G", 7)
                ws9 = workbook.create_sheet("H", 8)
                ws10 = workbook.create_sheet("I", 9)
                ws11 = workbook.create_sheet("J", 10)
                ws12 = workbook.create_sheet("CORE", 11)
                ws13 = workbook.create_sheet("Numbering", 12)
                ws14 = workbook.create_sheet("Extended_IUPAC_aa_code", 13)

                for rowdf in dataframe_to_rows(combined_df_results, index=False, 
                                                                    header=True):
                    ws1.append(rowdf)
                for cell in ws1[1]:
                        cell.style = 'Accent1'

                genotypes = ['A','B','C','D','E','F','G','H','I','J']
                worksheets = [ws2, ws3, ws4, ws5, ws6, ws7, 
                             ws8, ws9, ws10, ws11, ws12]
                header = ['Tube No.', 'Molis No.', 'Genotype', 'Surface',
                            'Polymerase (RT)', 'X region', 'Precore', 'Notes']
                for ws in worksheets:
                    ws.append(header)
                for rowdf in dataframe_to_rows(combined_df_results, 
                                               index=False, header=False):
                    for index, (gt, ws) in enumerate(zip(genotypes, worksheets)):
                        if gt in rowdf[2][0]: # Genotype first letter A,G, use np for NaN issue
                            ws.append(rowdf)
                    if rowdf[6]!= 'NA': #CORE
                        ws12.append(rowdf)
                for wsh in worksheets: 
                    for cell in wsh[1]:
                        cell.style = 'Accent2'

                # numbering worksheet
                acc_list = []
                target_list = []
                for s in spol_refseqs.values():
                    acc_number = s.split('_')[0]
                    target = s.split('_')[1:]
                    target =(' '.join(target)).replace('Genotype ','') #Genotype A
                    acc_list.append(acc_number)
                    target_list.append(target)

                acc_list.append(xprecore_refseq.id.split('_')[0])
                xprecore = (' '.join(xprecore_refseq.id.split('_')[1:]).
                                replace('Geno C WG','Xregion precore'))
                target_list.append(xprecore)
                data = {'Accession Number': acc_list, 
                        'Genotype and region for numbering': target_list}
                df_numbering = pd.DataFrame(data)

                for r in dataframe_to_rows(df_numbering, 
                                           index=False, header=True):
                    ws13.append(r)
                for cell in ws13[1]:
                        cell.style = 'Accent3'
                # amino acid codes worksheet
                aa_code_info = {'Symbol (combination)':['A','R','N','D','C','Q','E',
                                                        'G','H','I','M','L','K','F',
                                                        'P','S','T','W','Y','V','*',
                                                        'B (N/D)','X','Z (E/Q)',
                                                        'J (L/I)','U','O'],
                                'Amino acid':['Alanine', 'Arginine','Asparagine',
                                            'Aspartic acid','Cysteine', 'Glutamine',
                                            'Glutamic acid','Glycine','Histidine',
                                            'Isoleucine','Methionine','Leucine',
                                            'Lysine','Phenylalanine','Proline',
                                            'Serine','Threonine','Tryptophan',
                                            'Tyrosine','Valine','STOP',
                                            'Asparagine or Aspartic acid',
                                            'Unknown or other amino acid',
                                            'Glutamic acid or Glutamine',
                                            'Leucine or Isoleucine',
                                            'Selenocysteine','Pyrrolysine']
                                }
                df_amino_acid_code = pd.DataFrame(aa_code_info)

                for r in dataframe_to_rows(df_amino_acid_code, 
                                           index=False, header=True):
                    ws14.append(r)
                for cell in ws14[1]:
                        cell.style = 'Accent4'

                excel_file = path_to_run.joinpath(
                            f'Results_{run_name}_hbvma_v{__version__}.xlsx') 
                workbook.save(excel_file)
                shutil.rmtree(path_to_tmp) #Comment it out for debugging if needed 

                '''(F) SUMMARY OF INPUT/OUTPUT (LOG) AFTER ANALYSIS'''

                log_name = mu.summary_log(path_to_run, run_name, csv_seq_list, 
                                            xc_fasta_input, sp_fasta_input, 
                                            excel_file)
                print(f'''
    ANALYSIS COMPLETE
    Please check the details recorded in the log saved at 
    {log_name} 
    before looking at the results.

    Results are available at 
    {excel_file}.
    
    The alignments used to generate the excel report can be inspected at 
    {path_to_alns}.
    ''')

        except FileNotFoundError:
            print(f'''
    The file Sample_list.csv is not in the directory {path_to_input}.
    Please make sure to save it in the same directory where the FASTA sequences are.
    ''')

        except KeyError:
            print(f'''
    Please make sure:

    - the FASTA files have been provided in {path_to_input}
    - the header names in the Sample_list.csv in the directory 
      {path_to_input} are labelled as 'Seq ID' and 'Genotype'.
    - the FASTA sequences are labelled according to the convention
      tube_molisAmplicon (i.e. 01_H12345678sp)
    - the sequence id in the CSV file matches the identifier of the FASTA 
      files (either amplicon for the same sample). The name should be the 
      same after excluding sp/xc from the idenfifier of the FASTA file. 

            Example:        "CSV file"        
                        Seq ID              Genotype
                        01_H12345678         A1

                            "FASTA file"
            (correct)   >01_H12345678sp      
                        ATGTGTGGTGAT...

            (wrong)     >01_H12345678_sp      
                        ATGTGTGGTGAT...
            
            (wrong)     >1_H12345678sp      
                        ATGTGTGGTGAT...
                            ''')

    else:
        print(f'''
    The run {run_name} does NOT exist. 
    Make sure you enter the proper path to the run. 
        ''')

except FileNotFoundError:
    print(f'''
    Check if the path to the run directory {run_name} is the correct one. 
    {path_to_run}
    Runs usually should be named as E and a combination of numbers (i.e E1978).
     
    Please make sure a subdirectory Consensus is located in the directory {run_name}, 
    where the FASTA files to be analysed are saved.
    
    ''')
